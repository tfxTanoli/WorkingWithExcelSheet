import openpyxl as xl
from openpyxl.chart import Reference , BarChart
wb = xl.load_workbook('transaction.xlsx')
sheet = wb['Sheet1']

#count the rows of the spreadsheet
print(sheet.max_row)

#how to access the cells of the spreadsheet

cell = sheet['a1']
print(cell.value)

#How to access the rows and columns of spreadsheet using for loop
for i in range(2 , sheet.max_row+1):
    cell = sheet.cell(i , 3)
    # print(cell.value)
#how to add new column in the excel sheet
    correct_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(i , 4)
    corrected_price_cell.value = correct_price

values = Reference(sheet , min_row=2 , max_row=4 , min_col=4 ,max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart , 'j2')
wb.save('transactionSheet_with_barChart.xlsx')